from copy import deepcopy
from math import ceil
import numpy as np

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, colors, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

class Phenotype(object):
    def __init__(self, phenotypes, stats):
        self.phenotypes = phenotypes
        self.stats = stats

    def __getitem__(self, idx):
        return self.phenotypes[idx]

    def __len__(self):
        return len(self.phenotypes)


def interpret_phenotypes(*factors, item_idx2desc, threshold=None):
    phenotypes = []
    n_dims = len(factors)
    n_factors = factors[0].shape[1]
    item_sortidx = [np.argsort(-U, axis=0) for U in factors]

    if threshold is not None:
        factors = deepcopy(factors)
        for i, factor in enumerate(factors):
            factor[factor <= threshold[i]] = 0

    # get stats of phenotype overlapping
    def overlapping_stat(factor):
        positive_flag = factor > 0
        return positive_flag.sum(axis=1)

    stats = {
        'overlap': [overlapping_stat(factor) for factor in factors]
    }

    for r in range(n_factors):
        phenotype_definition = []
        for j in range(n_dims):
            dim_j = []
            for idx in item_sortidx[j][:, r]:
                if factors[j][idx, r] > 1e-4:
                    dim_j.append((idx, item_idx2desc[j][idx], factors[j][idx, r]))
                else:
                    break
            phenotype_definition.append(dim_j)
        phenotypes.append(phenotype_definition)
    return Phenotype(phenotypes, stats)


def coord(i, j):
    return '{}{}'.format(get_column_letter(j), i)


def phenotypes_to_excel_worksheet(phenotypes, dim_names, ws):
    n_dims = len(phenotypes[0])
    for i, pheno_r in enumerate(phenotypes):
        ws.merge_cells(coord(1, (n_dims+1)*i+1)+':'+coord(1, (n_dims+1)*i+n_dims))
        ws[coord(1, (n_dims+1)*i+1)] = 'Phenotype {:d}'.format(i+1)
        ws[coord(1, (n_dims+1)*i+1)].font = Font(bold=True)
        ws[coord(1, (n_dims+1)*i+1)].alignment = Alignment(horizontal='center', vertical='center')



        for j, name in enumerate(dim_names):
            ws[coord(2, (n_dims+1)*i+j+1)] = name
            ws[coord(2, (n_dims+1)*i+j+1)].alignment = Alignment(horizontal='center', vertical='center')
            ws[coord(2, (n_dims+1)*i+j+1)].font = Font(bold=True)
            for k, (idx, item, weight) in enumerate(pheno_r[j]):
                ws[coord(k+3, (n_dims+1)*i+j+1)] = '{} ({:.3f})'.format(item, weight)
                if phenotypes.stats['overlap'][j][idx] >= ceil(0.75 * len(phenotypes)):
                    ws.cell(row=k+3, column=(n_dims+1)*i+j+1).fill = PatternFill(fgColor='FF8B8B', fill_type='solid')
                elif phenotypes.stats['overlap'][j][idx] >= ceil(0.5 * len(phenotypes)):
                    ws.cell(row=k+3, column=(n_dims+1)*i+j+1).fill = PatternFill(fgColor='FFBF8B', fill_type='solid')
                elif phenotypes.stats['overlap'][j][idx] >= ceil(0.25 * len(phenotypes)):
                    ws.cell(row=k+3, column=(n_dims+1)*i+j+1).fill = PatternFill(fgColor='FFFF97', fill_type='solid')
            
            ws.column_dimensions[get_column_letter((n_dims+1)*i+j+1)].width = 50

                


def phenotypes_to_excel_file(phenotypes, dim_names, filepath, ws_name=None):
    wb = Workbook()
    ws = wb.active
    if ws_name:
        ws.title = ws_name
    phenotypes_to_excel_worksheet(phenotypes, dim_names, ws)
    wb.save(filepath)


def factor_matrices_to_excel(factors, item_dicts, filepath, threshold=None):
    dim_names, item_idx2desc = zip(*item_dicts)
    phenotypes = interpret_phenotypes(*factors, item_idx2desc=item_idx2desc, threshold=threshold)
    phenotypes_to_excel_file(phenotypes, dim_names, filepath)


def batch_factors_to_excel(factors_dict, item_dicts, filepath, threshold=None):
    wb = Workbook()
    wb.remove(wb.active)

    dim_names, item_idx2desc = zip(*item_dicts)

    for name, factors in factors_dict.items():
        ws = wb.create_sheet(name)
        phenotypes = interpret_phenotypes(*factors, item_idx2desc=item_idx2desc, threshold=threshold)
        phenotypes_to_excel_worksheet(phenotypes, dim_names, ws)
    wb.save(filepath)
