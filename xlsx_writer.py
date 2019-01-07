from math import ceil

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


def coord(i, j):
    return '{}{}'.format(get_column_letter(j), i)


def phenotypes_to_excel_worksheet(phenotypes, ws):
    n_dims = len(phenotypes[0])
    for i, pheno_r in enumerate(phenotypes):
        ws.merge_cells(coord(1, (n_dims + 1) * i + 1) + ':' + coord(1, (n_dims + 1) * i + n_dims))
        ws[coord(1, (n_dims + 1) * i + 1)] = 'Phenotype {:d}'.format(i + 1)
        ws[coord(1, (n_dims + 1) * i + 1)].font = Font(bold=True)
        ws[coord(1, (n_dims + 1) * i + 1)].alignment = Alignment(horizontal='center', vertical='center')

        for j, name in enumerate(phenotypes.dim_names):
            ws[coord(2, (n_dims + 1) * i + j + 1)] = name
            ws[coord(2, (n_dims + 1) * i + j + 1)].alignment = Alignment(horizontal='center', vertical='center')
            ws[coord(2, (n_dims + 1) * i + j + 1)].font = Font(bold=True)
            for k, (idx, item, weight) in enumerate(pheno_r[j]):
                ws[coord(k + 3, (n_dims + 1) * i + j + 1)] = '{} ({:.3f})'.format(item, weight)
                if phenotypes.stats['overlap'][j][idx] >= ceil(0.75 * len(phenotypes)):
                    ws.cell(row=k + 3, column=(n_dims + 1) * i + j + 1).fill = PatternFill(fgColor='FF8B8B',
                                                                                           fill_type='solid')
                elif phenotypes.stats['overlap'][j][idx] >= ceil(0.5 * len(phenotypes)):
                    ws.cell(row=k + 3, column=(n_dims + 1) * i + j + 1).fill = PatternFill(fgColor='FFBF8B',
                                                                                           fill_type='solid')
                elif phenotypes.stats['overlap'][j][idx] >= ceil(0.25 * len(phenotypes)):
                    ws.cell(row=k + 3, column=(n_dims + 1) * i + j + 1).fill = PatternFill(fgColor='FFFF97',
                                                                                           fill_type='solid')

            ws.column_dimensions[get_column_letter((n_dims + 1) * i + j + 1)].width = 50


def phenotypes_to_xlsx(phenotypes, filepath, ws_name=None):
    wb = Workbook()
    ws = wb.active
    if ws_name:
        ws.title = ws_name
    phenotypes_to_excel_worksheet(phenotypes, ws)
    wb.save(filepath)


def corrs_to_excel_worksheet(corrs_cols, ws):
    for dx_idx in range(len(corrs_cols)):
        col = corrs_cols[dx_idx]
        ws.merge_cells(coord(1, 3 * dx_idx + 1) + ':' + coord(1, 3 * dx_idx + 2))
        ws[coord(1, 3 * dx_idx + 1)] = col[0]
        ws[coord(1, 3 * dx_idx + 1)].font = Font(bold=True)
        ws[coord(1, 3 * dx_idx + 1)].alignment = Alignment(horizontal='center', vertical='center')
        ws.column_dimensions[get_column_letter(3 * dx_idx + 1)].width = 15
        ws.column_dimensions[get_column_letter(3 * dx_idx + 2)].width = 15
        ws.column_dimensions[get_column_letter(3 * dx_idx + 3)].width = 3

        for i, item in enumerate(col[1]):
            if item:
                ws[coord(i + 2, 3 * dx_idx + 1)] = item[0]
                ws[coord(i + 2, 3 * dx_idx + 2)] = item[1]
            else:
                break


def corrspondence_to_xlsx(corrs_cols, filepath, ws_name=None):
    wb = Workbook()
    ws = wb.active
    if ws_name:
        ws.title = ws_name
    corrs_to_excel_worksheet(corrs_cols, ws)
    wb.save(filepath)